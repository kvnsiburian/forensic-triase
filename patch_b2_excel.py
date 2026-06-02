"""
patch_b2_excel.py
=================
Patch B2: Ganti export CSV results menjadi Excel 5 sheet.

Sheet 1: PSList        — output windows.pslist
Sheet 2: PSTree        — output windows.pstree (flattened)
Sheet 3: NetScan       — output windows.netscan
Sheet 4: Malfind       — output windows.malware.malfind
Sheet 5: Klasifikasi   — hasil klasifikasi akhir (isi results.csv lama)

summary.csv tetap tidak berubah.

Cara pakai:
  cd ~/forensic_triase/platform
  python3 /tmp/patch_b2_excel.py
"""

from pathlib import Path
import shutil

BASE = Path.home() / "forensic_triase/platform"
REPORTER = BASE / "core/reporter.py"
MAIN     = BASE / "main.py"
APP      = BASE / "gui/app.py"

# ---------------------------------------------------------------------------
# Patch reporter.py
# ---------------------------------------------------------------------------

NEW_REPORTER = '''"""
reporter.py
===========
Modul ekspor hasil triase.

Output:
  1. [nama_dump]_[timestamp]_results.xlsx
     File Excel dengan 5 sheet:
       Sheet 1 — PSList    : output mentah windows.pslist
       Sheet 2 — PSTree    : output mentah windows.pstree (flattened)
       Sheet 3 — NetScan   : output mentah windows.netscan
       Sheet 4 — Malfind   : output mentah windows.malware.malfind
       Sheet 5 — Klasifikasi: hasil klasifikasi akhir per PID

  2. [nama_dump]_[timestamp]_summary.csv
     Satu baris per sesi — statistik ringkasan.

Author  : Kevin Armando Siburian (2221101800)
Program : Rekayasa Keamanan Siber - PSSN
"""

import csv
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

DEFAULT_OUTPUT_DIR = Path("/mnt/d/forensic_triase/output")


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _prepare_output_dir(output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    logger.info(f"Output folder: {output_dir}")


def _make_filename(dump_name: str, suffix: str, output_dir: Path, ext: str = "csv") -> Path:
    stem      = Path(dump_name).stem
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return output_dir / f"{stem}_{timestamp}_{suffix}.{ext}"


def _flatten_pstree(records: list) -> list:
    """Flatten nested pstree menjadi list datar."""
    flat = []
    for rec in records:
        node = {k: v for k, v in rec.items() if k != "__children"}
        flat.append(node)
        children = rec.get("__children", [])
        if children:
            flat.extend(_flatten_pstree(children))
    return flat


def _style_header(ws, row: int = 1):
    """Beri warna header biru gelap, teks putih bold."""
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")


def _style_row(ws, row: int, risk: str):
    """Warnai baris berdasarkan risk level."""
    colors = {
        "HIGH":   "FFDDD0",
        "MEDIUM": "FFF0CC",
        "LOW":    "FEFFD0",
        "CLEAN":  "D9F0D9",
    }
    fill_color = colors.get(risk, "FFFFFF")
    fill = PatternFill("solid", fgColor=fill_color)
    for cell in ws[row]:
        cell.fill = fill


def _auto_width(ws):
    """Set lebar kolom otomatis berdasarkan konten."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value else 0
                if val_len > max_len:
                    max_len = val_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _write_plugin_sheet(ws, records: list, sheet_title: str):
    """Tulis data plugin mentah ke worksheet."""
    ws.title = sheet_title

    if not records:
        ws.append(["(Tidak ada data)"])
        return

    # Ambil semua kolom dari record pertama
    headers = list(records[0].keys())
    ws.append(headers)
    _style_header(ws)

    for rec in records:
        row = [str(rec.get(h, "")) for h in headers]
        ws.append(row)

    _auto_width(ws)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Export results.xlsx — 5 sheet
# ---------------------------------------------------------------------------

def export_results_xlsx(
    classifications: list,
    plugin_results:  dict,
    dump_name: str,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
) -> Path:
    """
    Tulis hasil ke Excel dengan 5 sheet.

    Sheet 1-4: output mentah tiap plugin
    Sheet 5  : hasil klasifikasi akhir
    """
    _prepare_output_dir(output_dir)
    output_path = _make_filename(dump_name, "results", output_dir, ext="xlsx")

    wb = Workbook()

    # ── Sheet 1: PSList ───────────────────────────────────────────────
    ws1 = wb.active
    pslist = plugin_results.get("windows.pslist") or []
    _write_plugin_sheet(ws1, pslist, "PSList")

    # ── Sheet 2: PSTree (flattened) ───────────────────────────────────
    ws2 = wb.create_sheet()
    pstree_raw  = plugin_results.get("windows.pstree") or []
    pstree_flat = _flatten_pstree(pstree_raw)
    _write_plugin_sheet(ws2, pstree_flat, "PSTree")

    # ── Sheet 3: NetScan ──────────────────────────────────────────────
    ws3 = wb.create_sheet()
    netscan = plugin_results.get("windows.netscan") or []
    _write_plugin_sheet(ws3, netscan, "NetScan")

    # ── Sheet 4: Malfind ──────────────────────────────────────────────
    ws4 = wb.create_sheet()
    malfind = plugin_results.get("windows.malware.malfind") or []
    _write_plugin_sheet(ws4, malfind, "Malfind")

    # ── Sheet 5: Klasifikasi ──────────────────────────────────────────
    ws5 = wb.create_sheet(title="Klasifikasi")

    headers = [
        "PID", "Name", "Path", "Status", "Score", "Risk",
        "Rule1_Rogue", "Rule2_Network", "Rule3_Injection", "Rule4_ProcObj",
        "Reasons"
    ]
    ws5.append(headers)
    _style_header(ws5)

    for i, rec in enumerate(classifications, start=2):
        reasons_str = " | ".join(rec.get("Reasons", []))
        row = [
            rec["PID"],
            rec["Name"],
            rec["Path"] or "",
            rec["Status"],
            rec.get("Score", 0),
            rec.get("Risk", "CLEAN"),
            rec["Rule1_hit"],
            rec["Rule2_hit"],
            rec["Rule3_hit"],
            rec["Rule4_hit"],
            reasons_str,
        ]
        ws5.append(row)
        _style_row(ws5, i, rec.get("Risk", "CLEAN"))

    _auto_width(ws5)
    ws5.freeze_panes = "A2"

    wb.save(output_path)

    suspicious_count = sum(1 for r in classifications if r["Status"] == "SUSPICIOUS")
    logger.info(
        f"results.xlsx ditulis: {output_path.name} "
        f"({len(classifications)} baris, {suspicious_count} SUSPICIOUS, 5 sheet)"
    )
    return output_path


# ---------------------------------------------------------------------------
# Export summary.csv
# ---------------------------------------------------------------------------

def export_summary(
    classifications: list,
    dump_name: str,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
) -> Path:
    _prepare_output_dir(output_dir)
    output_path = _make_filename(dump_name, "summary", output_dir, ext="csv")

    total      = len(classifications)
    suspicious = sum(1 for r in classifications if r["Status"] == "SUSPICIOUS")
    clean      = total - suspicious
    high       = sum(1 for r in classifications if r.get("Risk") == "HIGH")
    medium     = sum(1 for r in classifications if r.get("Risk") == "MEDIUM")
    low        = sum(1 for r in classifications if r.get("Risk") == "LOW")
    rule1_hits = sum(1 for r in classifications if r["Rule1_hit"])
    rule2_hits = sum(1 for r in classifications if r["Rule2_hit"])
    rule3_hits = sum(1 for r in classifications if r["Rule3_hit"])
    rule4_hits = sum(1 for r in classifications if r["Rule4_hit"])

    fieldnames = [
        "Dump", "Timestamp", "Total_PID", "Suspicious",
        "High", "Medium", "Low", "Clean",
        "Rule1_Hits", "Rule2_Hits", "Rule3_Hits", "Rule4_Hits",
    ]

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerow({
            "Dump":       dump_name,
            "Timestamp":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Total_PID":  total,
            "Suspicious": suspicious,
            "High":       high,
            "Medium":     medium,
            "Low":        low,
            "Clean":      clean,
            "Rule1_Hits": rule1_hits,
            "Rule2_Hits": rule2_hits,
            "Rule3_Hits": rule3_hits,
            "Rule4_Hits": rule4_hits,
        })

    logger.info(f"summary.csv ditulis: {output_path.name}")
    return output_path


# ---------------------------------------------------------------------------
# Export all
# ---------------------------------------------------------------------------

def export_all(
    classifications: list,
    plugin_results:  dict,
    dump_name: str,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
) -> dict:
    """Export results.xlsx dan summary.csv sekaligus."""
    results_path = export_results_xlsx(
        classifications, plugin_results, dump_name, output_dir
    )
    summary_path = export_summary(classifications, dump_name, output_dir)
    return {
        "results_path": results_path,
        "summary_path": summary_path,
    }
'''

# ---------------------------------------------------------------------------
# Patch main.py — teruskan plugin_results ke export_all
# ---------------------------------------------------------------------------

OLD_MAIN_EXPORT = '        exported  = export_all(classifications, dump_name, output_dir)'
NEW_MAIN_EXPORT = '        exported  = export_all(classifications, plugin_results, dump_name, output_dir)'

# ---------------------------------------------------------------------------
# Patch app.py — update label dan referensi CSV ke Excel
# ---------------------------------------------------------------------------

OLD_APP_STATUS = (
    '        self._set_status(\n'
    '            f"Analisis selesai: {stats[\'total\']} PID | "\n'
    '            f"{stats[\'suspicious\']} SUSPICIOUS | {stats[\'clean\']} CLEAN  |  "\n'
    '            f"CSV: {result[\'results_path\'].name}"\n'
    '        )'
)
NEW_APP_STATUS = (
    '        self._set_status(\n'
    '            f"Analisis selesai: {stats[\'total\']} PID | "\n'
    '            f"{stats[\'suspicious\']} SUSPICIOUS | {stats[\'clean\']} CLEAN  |  "\n'
    '            f"Excel: {result[\'results_path\'].name}"\n'
    '        )'
)

OLD_APP_EXPORT_BTN = '        self._btn_export = tk.Button(\n            frame, text="💾  Export CSV",'
NEW_APP_EXPORT_BTN = '        self._btn_export = tk.Button(\n            frame, text="💾  Export Excel",'

OLD_APP_EXPORT_FN = '''    def _export_csv(self):
        """Export ulang CSV ke lokasi pilihan user."""
        if not self._classifications:
            return

        from core.reporter import export_all

        folder = filedialog.askdirectory(
            title="Pilih Folder Tujuan Export",
            initialdir="/mnt/d",
        )
        if not folder:
            return

        dump_name = Path(self._dump_path.get()).name
        exported  = export_all(
            self._classifications,
            dump_name,
            output_dir=Path(folder),
        )

        messagebox.showinfo(
            "Export Berhasil",
            f"File berhasil disimpan:\\n\\n"
            f"• {exported['results_path'].name}\\n"
            f"• {exported['summary_path'].name}\\n\\n"
            f"Lokasi: {folder}",
        )
        self._set_status(f"Export selesai: {folder}")'''

NEW_APP_EXPORT_FN = '''    def _export_csv(self):
        """Export ulang Excel ke lokasi pilihan user."""
        if not self._classifications:
            return

        from core.reporter import export_all

        folder = filedialog.askdirectory(
            title="Pilih Folder Tujuan Export",
            initialdir="/mnt/d",
        )
        if not folder:
            return

        dump_name    = Path(self._dump_path.get()).name
        plugin_results = getattr(self, "_plugin_results", {})
        exported  = export_all(
            self._classifications,
            plugin_results,
            dump_name,
            output_dir=Path(folder),
        )

        messagebox.showinfo(
            "Export Berhasil",
            f"File berhasil disimpan:\\n\\n"
            f"• {exported['results_path'].name}\\n"
            f"• {exported['summary_path'].name}\\n\\n"
            f"Lokasi: {folder}",
        )
        self._set_status(f"Export selesai: {folder}")'''

# Simpan plugin_results ke self setelah analisis selesai
OLD_APP_DONE = '        self._classifications = result["classifications"]'
NEW_APP_DONE = (
    '        self._classifications  = result["classifications"]\n'
    '        self._plugin_results   = result.get("plugin_results", {})'
)

# main.py juga perlu return plugin_results
OLD_MAIN_RESULT = '''        result.update({
            "success":         True,
            "classifications": classifications,
            "results_path":    exported["results_path"],
            "summary_path":    exported["summary_path"],
            "stats": {
                "total":      total,
                "suspicious": suspicious,
                "clean":      clean,
            },
        })'''

NEW_MAIN_RESULT = '''        result.update({
            "success":         True,
            "classifications": classifications,
            "plugin_results":  plugin_results,
            "results_path":    exported["results_path"],
            "summary_path":    exported["summary_path"],
            "stats": {
                "total":      total,
                "suspicious": suspicious,
                "clean":      clean,
            },
        })'''


# ---------------------------------------------------------------------------
# Helper patch
# ---------------------------------------------------------------------------

def patch_file(path: Path, replacements: list, label: str):
    shutil.copy(path, str(path) + ".bak2")
    content = path.read_text(encoding="utf-8")
    for i, (old, new) in enumerate(replacements, 1):
        if old not in content:
            print(f"  [!] Patch {i} GAGAL — string tidak ditemukan di {label}")
            continue
        content = content.replace(old, new, 1)
        print(f"  [✓] Patch {i} berhasil")
    path.write_text(content, encoding="utf-8")


# ---------------------------------------------------------------------------
# Jalankan semua patch
# ---------------------------------------------------------------------------

print("\n" + "=" * 60)
print("  PATCH B2: Export Excel 5 Sheet")
print("=" * 60)

print("\n[1/3] Menulis ulang reporter.py...")
shutil.copy(REPORTER, str(REPORTER) + ".bak2")
REPORTER.write_text(NEW_REPORTER, encoding="utf-8")
print("  [✓] reporter.py ditulis ulang")

print("\n[2/3] Patching main.py...")
patch_file(MAIN, [
    (OLD_MAIN_EXPORT, NEW_MAIN_EXPORT),
    (OLD_MAIN_RESULT, NEW_MAIN_RESULT),
], "main.py")

print("\n[3/3] Patching app.py...")
patch_file(APP, [
    (OLD_APP_STATUS,     NEW_APP_STATUS),
    (OLD_APP_EXPORT_BTN, NEW_APP_EXPORT_BTN),
    (OLD_APP_EXPORT_FN,  NEW_APP_EXPORT_FN),
    (OLD_APP_DONE,       NEW_APP_DONE),
], "app.py")

print("\n" + "=" * 60)
print("  SELESAI.")
print("  Test:")
print("  python3 ~/forensic_triase/platform/main.py /mnt/d/forensic_triase/dataset/infected_rogue.raw")
print("=" * 60 + "\n")