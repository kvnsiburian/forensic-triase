"""
reporter.py
===========
Modul ekspor hasil triase.

Platform menghasilkan dua jenis berkas keluaran sekaligus:

A. Berkas Excel (.xlsx) — untuk keterbacaan
  1. [nama_dump]_[timestamp]_results.xlsx
     File Excel dengan 7 sheet:
       Sheet 1 — PSList    : output mentah windows.pslist
       Sheet 2 — PSTree    : output mentah windows.pstree (flattened)
       Sheet 3 — NetScan   : output mentah windows.netscan
       Sheet 4 — Malfind   : output mentah windows.malware.malfind
       Sheet 5 — DllList   : output mentah windows.dlllist
       Sheet 6 — Handles   : output mentah windows.handles
       Sheet 7 — Klasifikasi: hasil klasifikasi akhir per PID
  2. [nama_dump]_[timestamp]_summary.xlsx
     Satu baris per sesi — statistik ringkasan.

B. Berkas CSV (.csv) — agar sesuai dengan alur kerja Tim LFD
  1. [nama_dump]_[timestamp]_klasifikasi.csv : hasil klasifikasi akhir per PID
  2. [nama_dump]_[timestamp]_summary.csv     : statistik ringkasan satu baris
  3. [nama_dump]_[timestamp]_[plugin].csv    : output mentah tiap plugin

Author  : Kevin Armando Siburian (2221101800)
Program : Rekayasa Keamanan Siber - PSSN
"""

import csv
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def _resolve_default_output_dir() -> Path:
    """Tentukan folder output default yang selalu tersedia.

    Utamakan drive D: (lewat mount /mnt/d pada lingkungan WSL) bila ada, supaya
    hasil mudah diakses dari Windows. Jika drive itu tidak ada (misalnya laptop
    tanpa D: atau bukan WSL), gunakan folder di home pengguna yang pasti bisa
    ditulisi, sehingga hasil tidak tersimpan di lokasi yang membingungkan.
    """
    d_drive = Path("/mnt/d")
    if d_drive.exists():
        return d_drive / "forensic_triase" / "output"
    return Path.home() / "forensic_triase_output"


DEFAULT_OUTPUT_DIR = _resolve_default_output_dir()


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _prepare_output_dir(output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    logger.info(f"Output folder: {output_dir}")


def _make_filename(dump_name: str, suffix: str, output_dir: Path, ext: str = "xlsx") -> Path:
    stem      = Path(dump_name).stem
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return output_dir / f"{stem}_{timestamp}_{suffix}.{ext}"


def _clean_cell(value) -> str:
    """Buang karakter kontrol yang ditolak Excel (muncul di output
    windows.handles pada nama objek kernel tertentu)."""
    return ILLEGAL_CHARACTERS_RE.sub("", str(value))


def _flatten_pstree(records: list) -> list:
    """Flatten nested pstree menjadi list datar."""
    flat = []
    for rec in records:
        node = {k: v for k, v in rec.items() if k != "__children"}
        flat.append(node)
        children = rec.get("__children", [])
        if children:
            flat.extend(_flatten_pstree(children))
    return flat


def _style_header(ws, row: int = 1):
    """Beri warna header biru gelap, teks putih bold."""
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")


def _style_row(ws, row: int, status: str):
    """
    Warnai baris berdasarkan Status (binary).
    Per arahan Pak Rahmat (3 Juni 2026): scoring layer dihapus,
    pewarnaan disederhanakan menjadi 2 kategori saja.
    """
    colors = {
        "SUSPICIOUS": "FFE0CC",  # oranye muda
        "CLEAN":      "D9F0D9",  # hijau muda
    }
    fill_color = colors.get(status, "FFFFFF")
    fill = PatternFill("solid", fgColor=fill_color)
    for cell in ws[row]:
        cell.fill = fill


def _auto_width(ws):
    """Set lebar kolom otomatis berdasarkan konten."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value else 0
                if val_len > max_len:
                    max_len = val_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


# Kolom yang di-hide per sheet
HIDE_COLUMNS = {
    "PSList":  {"Handles", "__children", "File output"},
    "PSTree":  {"Handles", "__children"},
    "NetScan": {"__children"},
    "Malfind": {"__children"},
    "DllList": {"__children", "File output"},
    "Handles": {"__children"},
}


def _write_plugin_sheet(ws, records: list, sheet_title: str):
    """Tulis data plugin mentah ke worksheet, hide kolom yang tidak perlu."""
    ws.title = sheet_title

    if not records:
        ws.append(["(Tidak ada data)"])
        return

    # Ambil semua kolom dari record pertama
    headers = list(records[0].keys())
    ws.append(headers)
    _style_header(ws)

    for rec in records:
        row = [_clean_cell(rec.get(h, "")) for h in headers]
        ws.append(row)

    _auto_width(ws)
    ws.freeze_panes = "A2"

    # Hide kolom yang tidak perlu
    hidden = HIDE_COLUMNS.get(sheet_title, set())
    for col_idx, header in enumerate(headers, start=1):
        if header in hidden:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].hidden = True


# ---------------------------------------------------------------------------
# Export results.xlsx — 7 sheet
# ---------------------------------------------------------------------------

def export_results_xlsx(
    classifications: list,
    plugin_results:  dict,
    dump_name: str,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
) -> Path:
    """
    Tulis hasil ke Excel dengan 7 sheet.

    Sheet 1-6: output mentah tiap plugin
    Sheet 7  : hasil klasifikasi akhir
    """
    _prepare_output_dir(output_dir)
    output_path = _make_filename(dump_name, "results", output_dir, ext="xlsx")

    wb = Workbook()

    # ── Sheet 1: PSList ───────────────────────────────────────────────
    ws1 = wb.active
    pslist = plugin_results.get("windows.pslist") or []
    _write_plugin_sheet(ws1, pslist, "PSList")

    # ── Sheet 2: PSTree (flattened) ───────────────────────────────────
    ws2 = wb.create_sheet()
    pstree_raw  = plugin_results.get("windows.pstree") or []
    pstree_flat = _flatten_pstree(pstree_raw)
    _write_plugin_sheet(ws2, pstree_flat, "PSTree")

    # ── Sheet 3: NetScan ──────────────────────────────────────────────
    ws3 = wb.create_sheet()
    netscan = plugin_results.get("windows.netscan") or []
    _write_plugin_sheet(ws3, netscan, "NetScan")

    # ── Sheet 4: Malfind ──────────────────────────────────────────────
    ws4 = wb.create_sheet()
    malfind = plugin_results.get("windows.malware.malfind") or []
    _write_plugin_sheet(ws4, malfind, "Malfind")

    # ── Sheet 5: DllList ──────────────────────────────────────────────
    ws5 = wb.create_sheet()
    dlllist = plugin_results.get("windows.dlllist") or []
    _write_plugin_sheet(ws5, dlllist, "DllList")

    # ── Sheet 6: Handles ──────────────────────────────────────────────
    ws6 = wb.create_sheet()
    handles = plugin_results.get("windows.handles") or []
    _write_plugin_sheet(ws6, handles, "Handles")

    # ── Sheet 7: Klasifikasi ──────────────────────────────────────────
    # Per arahan Pak Rahmat (3 Juni 2026): kolom Score dan Risk dihapus dari
    # output Excel — klasifikasi binary murni. Logika scoring dihapus total
    # dari analyzer.py per arahan Pak Rahmat (3 Juni 2026) — klasifikasi binary murni.
    ws7 = wb.create_sheet(title="Klasifikasi")

    headers = [
        "PID", "Name", "Path", "Status",
        "Rule1_Rogue", "Rule2_Network", "Rule3_Injection", "Rule4_ProcObj",
        "Reasons"
    ]
    ws7.append(headers)
    _style_header(ws7)

    for i, rec in enumerate(classifications, start=2):
        reasons_str = _clean_cell(" | ".join(rec.get("Reasons", [])))
        row = [
            rec["PID"],
            rec["Name"],
            rec["Path"] or "",
            rec["Status"],
            rec["Rule1_hit"],
            rec["Rule2_hit"],
            rec["Rule3_hit"],
            rec["Rule4_hit"],
            reasons_str,
        ]
        ws7.append(row)
        _style_row(ws7, i, rec.get("Status", "CLEAN"))

    _auto_width(ws7)
    ws7.freeze_panes = "A2"

    wb.save(output_path)

    suspicious_count = sum(1 for r in classifications if r["Status"] == "SUSPICIOUS")
    logger.info(
        f"results.xlsx ditulis: {output_path.name} "
        f"({len(classifications)} baris, {suspicious_count} SUSPICIOUS, 7 sheet)"
    )
    return output_path


# ---------------------------------------------------------------------------
# Export summary.xlsx
# ---------------------------------------------------------------------------

def export_summary(
    classifications: list,
    dump_name: str,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
) -> Path:
    _prepare_output_dir(output_dir)
    output_path = _make_filename(dump_name, "summary", output_dir, ext="xlsx")

    total      = len(classifications)
    suspicious = sum(1 for r in classifications if r["Status"] == "SUSPICIOUS")
    clean      = total - suspicious
    # Per arahan Pak Rahmat (3 Juni 2026): kolom High/Medium/Low dihapus
    # dari summary — klasifikasi binary murni.
    rule1_hits = sum(1 for r in classifications if r["Rule1_hit"])
    rule2_hits = sum(1 for r in classifications if r["Rule2_hit"])
    rule3_hits = sum(1 for r in classifications if r["Rule3_hit"])
    rule4_hits = sum(1 for r in classifications if r["Rule4_hit"])

    fieldnames = [
        "Dump", "Timestamp", "Total_PID", "Suspicious", "Clean",
        "Rule1_Hits", "Rule2_Hits", "Rule3_Hits", "Rule4_Hits",
    ]

    values = [
        dump_name,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        total, suspicious, clean,
        rule1_hits, rule2_hits, rule3_hits, rule4_hits,
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(fieldnames)
    _style_header(ws)
    ws.append(values)
    _auto_width(ws)
    wb.save(output_path)

    logger.info(f"summary.xlsx ditulis: {output_path.name}")
    return output_path


# ---------------------------------------------------------------------------
# Export CSV — versi datar dari isi Excel
# ---------------------------------------------------------------------------

def _write_csv(path: Path, headers: list, rows: list) -> None:
    """Tulis satu tabel ke berkas CSV (UTF-8 dengan BOM agar rapi di Excel)."""
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


def export_klasifikasi_csv(
    classifications: list,
    dump_name: str,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
) -> Path:
    """Tulis tabel klasifikasi akhir per PID ke CSV.

    Isinya sama dengan sheet Klasifikasi pada berkas Excel, hanya saja datar
    tanpa pewarnaan. Ini berkas utama yang dipakai analis untuk menindaklanjuti.
    """
    _prepare_output_dir(output_dir)
    output_path = _make_filename(dump_name, "klasifikasi", output_dir, ext="csv")

    headers = [
        "PID", "Name", "Path", "Status",
        "Rule1_Rogue", "Rule2_Network", "Rule3_Injection", "Rule4_ProcObj",
        "Reasons",
    ]
    rows = []
    for rec in classifications:
        rows.append([
            rec["PID"],
            rec["Name"],
            rec["Path"] or "",
            rec["Status"],
            rec["Rule1_hit"],
            rec["Rule2_hit"],
            rec["Rule3_hit"],
            rec["Rule4_hit"],
            _clean_cell(" | ".join(rec.get("Reasons", []))),
        ])
    _write_csv(output_path, headers, rows)

    logger.info(f"klasifikasi.csv ditulis: {output_path.name} ({len(rows)} baris)")
    return output_path


# ---------------------------------------------------------------------------
# Export all
# ---------------------------------------------------------------------------

def export_all(
    classifications: list,
    plugin_results:  dict,
    dump_name: str,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
) -> dict:
    """Export tiga berkas keluaran:

      1. [dump]_results.xlsx     : Excel lengkap 7 sheet (data mentah + klasifikasi)
      2. [dump]_klasifikasi.csv  : tabel klasifikasi/vonis (sama dgn sheet Klasifikasi)
      3. [dump]_summary.xlsx     : ringkasan statistik singkat (Excel saja)

    Excel untuk keterbacaan, CSV untuk kesesuaian dengan alur kerja Tim LFD.
    """
    # ── Excel ──────────────────────────────────────────────────────────
    results_path = export_results_xlsx(
        classifications, plugin_results, dump_name, output_dir
    )
    summary_path = export_summary(classifications, dump_name, output_dir)

    # ── CSV (hanya tabel klasifikasi/vonis) ────────────────────────────
    klasifikasi_csv = export_klasifikasi_csv(classifications, dump_name, output_dir)

    return {
        "results_path":     results_path,      # Excel lengkap
        "summary_path":     summary_path,      # Excel ringkasan
        "klasifikasi_csv":  klasifikasi_csv,   # CSV vonis
    }